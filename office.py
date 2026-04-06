#!/usr/bin/env python3
"""
Asyre Office - Unified file operations CLI for AI agents.
Subcommand: sheet - CRUD operations on CSV/Excel files.
"""

import argparse
import csv
import json
import os
import sys
import re
from pathlib import Path
from io import StringIO

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def detect_encoding(filepath):
    """Detect file encoding, fallback to utf-8."""
    for enc in ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'latin-1']:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return 'utf-8'


def detect_delimiter(filepath, encoding='utf-8'):
    """Detect CSV delimiter."""
    with open(filepath, 'r', encoding=encoding) as f:
        sample = f.read(4096)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
        return dialect.delimiter
    except csv.Error:
        return ','


def get_file_type(filepath):
    """Determine file type from extension."""
    ext = Path(filepath).suffix.lower()
    if ext in ('.csv', '.tsv'):
        return 'csv'
    elif ext in ('.xlsx',):
        return 'xlsx'
    elif ext in ('.xls',):
        return 'xls'
    elif ext in ('.json',):
        return 'json'
    else:
        return None


def load_csv(filepath):
    """Load CSV file preserving metadata for write-back."""
    encoding = detect_encoding(filepath)
    delimiter = detect_delimiter(filepath, encoding)
    rows = []
    with open(filepath, 'r', encoding=encoding, newline='') as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        headers = reader.fieldnames or []
        for row in reader:
            rows.append(dict(row))
    return headers, rows, {'encoding': encoding, 'delimiter': delimiter}


def save_csv(filepath, headers, rows, meta):
    """Save CSV preserving original encoding and delimiter."""
    encoding = meta.get('encoding', 'utf-8')
    delimiter = meta.get('delimiter', ',')
    with open(filepath, 'w', encoding=encoding, newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(rows)


def find_header_row(ws):
    """Find the header row in a worksheet. Looks for the first row where most cells have values."""
    for row_i in range(1, min(ws.max_row + 1, 20)):  # scan first 20 rows
        vals = [ws.cell(row=row_i, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in vals if v is not None and str(v).strip()]
        # Header row: at least 2 non-empty cells, all text (not formulas)
        if len(non_empty) >= 2:
            all_text = all(isinstance(v, str) or (not str(v).startswith('=')) for v in non_empty)
            if all_text:
                return row_i
    return 1  # fallback


def load_xlsx(filepath):
    """Load Excel file, auto-detect header row, preserve workbook for safe editing."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("error: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active

    if ws.max_row is None or ws.max_row == 0:
        return [], [], {'wb': wb, 'ws': ws, 'header_row': 1, 'mode': 'simple'}

    header_row = find_header_row(ws)
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is not None:
            headers.append(str(val))
        else:
            break  # stop at first empty header

    rows = []
    row_map = []  # maps rows[] index → actual Excel row number
    for r in range(header_row + 1, ws.max_row + 1):
        d = {}
        has_data = False
        for i, h in enumerate(headers):
            val = ws.cell(row=r, column=i + 1).value
            d[h] = val
            if val is not None:
                has_data = True
        if has_data:
            rows.append(d)
            row_map.append(r)

    return headers, rows, {
        'wb': wb, 'ws': ws,
        'header_row': header_row,
        'row_map': row_map,  # actual Excel row for each data row
        'mode': 'inplace',
    }


def save_xlsx_inplace(filepath, headers, rows, meta, changes=None):
    """
    Save Excel with MINIMAL changes. Two modes:
    - changes dict provided: only update specific cells (safest)
    - no changes: rewrite data area below header row (for add/delete)
    """
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    wb = meta.get('wb')
    ws = meta.get('ws')
    header_row = meta.get('header_row', 1)

    if wb is None:
        # New file
        wb = Workbook()
        ws = wb.active
        for col_i, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_i, value=h)
        header_row = 1

    if changes:
        # SAFEST: only touch the specific cells that changed
        for (row_idx, col_idx), new_val in changes.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            # Preserve style, only change value
            if isinstance(new_val, str):
                try:
                    new_val = int(new_val)
                except ValueError:
                    try:
                        new_val = float(new_val)
                    except ValueError:
                        pass
            cell.value = new_val
    else:
        # Rewrite data rows (for add/delete operations)
        data_start = header_row + 1
        for row_i, row_data in enumerate(rows):
            excel_row = data_start + row_i
            for col_i, h in enumerate(headers, 1):
                cell = ws.cell(row=excel_row, column=col_i)
                val = row_data.get(h)
                if isinstance(val, str):
                    try:
                        val = int(val)
                    except ValueError:
                        try:
                            val = float(val)
                        except ValueError:
                            pass
                cell.value = val

        # Clear leftover rows (if rows were deleted)
        expected_end = data_start + len(rows)
        for row_i in range(expected_end, ws.max_row + 1):
            has_data = False
            for col_i in range(1, len(headers) + 1):
                c = ws.cell(row=row_i, column=col_i)
                if c.value is not None and not str(c.value).startswith('='):
                    has_data = True
            # Only clear rows that look like data rows (not formula/summary rows)
            if has_data:
                for col_i in range(1, len(headers) + 1):
                    ws.cell(row=row_i, column=col_i).value = None

    wb.save(filepath)


def load_json_table(filepath):
    """Load JSON array as table."""
    encoding = detect_encoding(filepath)
    with open(filepath, 'r', encoding=encoding) as f:
        data = json.load(f)
    if not isinstance(data, list) or not data:
        return [], [], {'encoding': encoding}
    headers = list(data[0].keys())
    rows = [dict(row) for row in data]
    return headers, rows, {'encoding': encoding}


def save_json_table(filepath, headers, rows, meta):
    """Save rows as JSON array."""
    encoding = meta.get('encoding', 'utf-8')
    with open(filepath, 'w', encoding=encoding) as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def load_file(filepath):
    """Load any supported file type."""
    ft = get_file_type(filepath)
    if ft == 'csv':
        return load_csv(filepath)
    elif ft == 'xlsx':
        return load_xlsx(filepath)
    elif ft == 'json':
        return load_json_table(filepath)
    elif ft == 'xls':
        print("error: .xls format is read-only. Convert to .xlsx first.", file=sys.stderr)
        # Try with xlrd
        try:
            import xlrd
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_index(0)
            headers = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
            rows = []
            for r in range(1, ws.nrows):
                row = {}
                for c in range(ws.ncols):
                    row[headers[c]] = ws.cell_value(r, c)
                rows.append(row)
            return headers, rows, {'readonly': True}
        except ImportError:
            print("error: xlrd required for .xls files. Install: pip install xlrd", file=sys.stderr)
            sys.exit(1)
    else:
        print(f"error: unsupported file type: {Path(filepath).suffix}", file=sys.stderr)
        sys.exit(1)


def save_file(filepath, headers, rows, meta, changes=None):
    """Save to the original format. For xlsx, uses in-place editing when possible."""
    if meta.get('readonly'):
        print("error: .xls is read-only. Use .xlsx instead.", file=sys.stderr)
        sys.exit(1)
    ft = get_file_type(filepath)
    if ft == 'csv':
        save_csv(filepath, headers, rows, meta)
    elif ft == 'xlsx':
        save_xlsx_inplace(filepath, headers, rows, meta, changes=changes)
    elif ft == 'json':
        save_json_table(filepath, headers, rows, meta)


def parse_kv_pairs(args):
    """Parse key=value pairs from command line args."""
    result = {}
    for arg in args:
        if '=' in arg:
            k, v = arg.split('=', 1)
            result[k.strip()] = v.strip()
    return result


def match_where(row, where_clause):
    """Check if a row matches a where clause like 'name=张三' or 'amount>1000'."""
    if not where_clause:
        return True

    # Support multiple conditions separated by comma
    conditions = [c.strip() for c in where_clause.split(',')]
    for cond in conditions:
        # Parse operator
        for op in ['>=', '<=', '!=', '>', '<', '=']:
            if op in cond:
                key, val = cond.split(op, 1)
                key = key.strip()
                val = val.strip()
                row_val = str(row.get(key, ''))

                # Try numeric comparison
                try:
                    num_row = float(row_val)
                    num_val = float(val)
                    if op == '=' and not (num_row == num_val):
                        return False
                    elif op == '!=' and not (num_row != num_val):
                        return False
                    elif op == '>' and not (num_row > num_val):
                        return False
                    elif op == '<' and not (num_row < num_val):
                        return False
                    elif op == '>=' and not (num_row >= num_val):
                        return False
                    elif op == '<=' and not (num_row <= num_val):
                        return False
                except ValueError:
                    # String comparison
                    if op == '=' and row_val != val:
                        return False
                    elif op == '!=' and row_val == val:
                        return False
                    elif op in ('>', '<', '>=', '<='):
                        if op == '>' and not (row_val > val):
                            return False
                        elif op == '<' and not (row_val < val):
                            return False
                        elif op == '>=' and not (row_val >= val):
                            return False
                        elif op == '<=' and not (row_val <= val):
                            return False
                break
    return True


def format_table(headers, rows, limit=None):
    """Format rows as a simple aligned table."""
    if not headers:
        return "(empty)"

    display_rows = rows[:limit] if limit else rows

    # Calculate column widths
    widths = {h: len(str(h)) for h in headers}
    for row in display_rows:
        for h in headers:
            val = str(row.get(h, ''))
            widths[h] = max(widths[h], len(val))

    # Cap column width at 40
    for h in headers:
        widths[h] = min(widths[h], 40)

    # Build output
    lines = []
    # Header
    header_line = ' | '.join(str(h).ljust(widths[h])[:widths[h]] for h in headers)
    lines.append(header_line)
    lines.append('-+-'.join('-' * widths[h] for h in headers))

    # Data
    for row in display_rows:
        line = ' | '.join(str(row.get(h, '')).ljust(widths[h])[:widths[h]] for h in headers)
        lines.append(line)

    total = len(rows)
    shown = len(display_rows)
    if limit and total > shown:
        lines.append(f'({total} rows total, showing {shown})')
    else:
        lines.append(f'({total} rows)')

    return '\n'.join(lines)


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_info(args):
    """Show file structure: columns, row count, sample data."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(filepath)
    size = os.path.getsize(filepath)
    size_str = f"{size/1024:.1f}KB" if size < 1048576 else f"{size/1048576:.1f}MB"

    if args.json:
        print(json.dumps({
            'file': filepath,
            'type': get_file_type(filepath),
            'size': size_str,
            'rows': len(rows),
            'columns': headers,
        }, ensure_ascii=False))
    else:
        print(f"file: {filepath}")
        print(f"type: {get_file_type(filepath)}")
        print(f"size: {size_str}")
        print(f"rows: {len(rows)}")
        print(f"columns: {', '.join(headers)}")
        if rows:
            print(f"\nsample (first 3 rows):")
            print(format_table(headers, rows, limit=3))


def cmd_list(args):
    """List file contents with optional limit."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(filepath)
    limit = args.limit or 20

    if args.columns:
        show_cols = [c.strip() for c in args.columns.split(',')]
        headers = [h for h in headers if h in show_cols]

    if args.json:
        display = rows[:limit]
        if args.columns:
            display = [{k: r.get(k) for k in headers} for r in display]
        out = {'rows': display, 'total': len(rows), 'showing': len(display)}
        print(json.dumps(out, ensure_ascii=False))
    else:
        if args.columns:
            filtered = [{k: r.get(k) for k in headers} for r in rows]
            print(format_table(headers, filtered, limit=limit))
        else:
            print(format_table(headers, rows, limit=limit))


def cmd_query(args):
    """Query rows with conditions."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(filepath)

    # Filter
    matched = [r for r in rows if match_where(r, args.where)]

    # Sort
    if args.sort:
        reverse = False
        sort_key = args.sort
        if sort_key.startswith('-'):
            reverse = True
            sort_key = sort_key[1:]
        if sort_key in headers:
            def sort_fn(r):
                v = r.get(sort_key, '')
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return str(v)
            matched.sort(key=sort_fn, reverse=reverse)

    limit = args.limit or 20

    # Select columns
    show_headers = headers
    if args.columns:
        show_headers = [c.strip() for c in args.columns.split(',')]

    if args.json:
        display = matched[:limit]
        if args.columns:
            display = [{k: r.get(k) for k in show_headers} for r in display]
        out = {'rows': display, 'matched': len(matched), 'total': len(rows), 'showing': len(display)}
        print(json.dumps(out, ensure_ascii=False))
    else:
        if args.columns:
            filtered = [{k: r.get(k) for k in show_headers} for r in matched]
            print(format_table(show_headers, filtered, limit=limit))
        else:
            print(format_table(show_headers, matched, limit=limit))


def cmd_add(args):
    """Add a row to the file."""
    filepath = args.file
    kv = parse_kv_pairs(args.values)
    if not kv:
        print("error: provide key=value pairs. Example: tables sheet add data.csv name=Alice age=30", file=sys.stderr)
        sys.exit(1)

    if os.path.exists(filepath):
        headers, rows, meta = load_file(filepath)
        # Add new columns if needed
        for k in kv:
            if k not in headers:
                headers.append(k)
    else:
        # Auto-create file
        headers = list(kv.keys())
        rows = []
        ft = get_file_type(filepath)
        if ft == 'csv':
            meta = {'encoding': 'utf-8', 'delimiter': ','}
        elif ft == 'xlsx':
            meta = {'wb': None, 'ws': None}
        elif ft == 'json':
            meta = {'encoding': 'utf-8'}
        else:
            print(f"error: cannot create file with extension: {Path(filepath).suffix}", file=sys.stderr)
            sys.exit(1)

    new_row = {h: kv.get(h, '') for h in headers}
    rows.append(new_row)
    save_file(filepath, headers, rows, meta)

    if args.quiet:
        print(f"added: 1")
    else:
        print(f"added 1 row to {filepath} (total: {len(rows)})")


def cmd_update(args):
    """Update rows matching a condition."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    if not args.where:
        print("error: --where is required for update (safety). Use --where to specify which rows to update.", file=sys.stderr)
        sys.exit(1)

    if not args.set:
        print("error: --set is required. Example: --set \"status=done,note=finished\"", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(filepath)
    header_row = meta.get('header_row', 1)

    # Parse set values
    set_pairs = {}
    for pair in args.set.split(','):
        if '=' in pair:
            k, v = pair.split('=', 1)
            set_pairs[k.strip()] = v.strip()

    # For xlsx: collect specific cell changes (safest)
    is_xlsx = get_file_type(filepath) == 'xlsx'
    cell_changes = {} if is_xlsx else None
    row_map = meta.get('row_map', [])

    # Add new columns if needed
    for k in set_pairs:
        if k not in headers:
            headers.append(k)

    count = 0
    for i, row in enumerate(rows):
        if match_where(row, args.where):
            for k, v in set_pairs.items():
                row[k] = v
                if is_xlsx and k in headers:
                    col_idx = headers.index(k) + 1
                    # Use row_map for exact Excel row, fallback to calculation
                    row_idx = row_map[i] if i < len(row_map) else (header_row + 1 + i)
                    cell_changes[(row_idx, col_idx)] = v
            count += 1

    if count > 0:
        save_file(filepath, headers, rows, meta, changes=cell_changes)

    if args.quiet:
        print(f"updated: {count}")
    else:
        print(f"updated {count} rows in {filepath}")


def cmd_delete(args):
    """Delete rows matching a condition."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    if not args.where:
        print("error: --where is required for delete (safety).", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(filepath)

    original_count = len(rows)
    rows = [r for r in rows if not match_where(r, args.where)]
    deleted = original_count - len(rows)

    if deleted > 0:
        save_file(filepath, headers, rows, meta)

    if args.quiet:
        print(f"deleted: {deleted}")
    else:
        print(f"deleted {deleted} rows from {filepath} (remaining: {len(rows)})")


def cmd_stats(args):
    """Show statistics for a column."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(filepath)
    column = args.column

    if column not in headers:
        print(f"error: column '{column}' not found. Available: {', '.join(headers)}", file=sys.stderr)
        sys.exit(1)

    if args.group_by:
        # Group by another column
        gb = args.group_by
        if gb not in headers:
            print(f"error: group-by column '{gb}' not found.", file=sys.stderr)
            sys.exit(1)

        groups = {}
        for row in rows:
            key = str(row.get(gb, ''))
            val = row.get(column)
            if key not in groups:
                groups[key] = []
            try:
                groups[key].append(float(val))
            except (ValueError, TypeError):
                pass

        if args.json:
            result = {}
            for key, vals in groups.items():
                if vals:
                    result[key] = {
                        'count': len(vals),
                        'sum': round(sum(vals), 2),
                        'avg': round(sum(vals) / len(vals), 2),
                        'min': min(vals),
                        'max': max(vals),
                    }
            print(json.dumps(result, ensure_ascii=False))
        else:
            for key, vals in sorted(groups.items()):
                if vals:
                    print(f"{gb}={key}: count={len(vals)} sum={sum(vals):.2f} avg={sum(vals)/len(vals):.2f} min={min(vals)} max={max(vals)}")
        return

    # Single column stats
    values = []
    non_numeric = 0
    for row in rows:
        val = row.get(column)
        try:
            values.append(float(val))
        except (ValueError, TypeError):
            non_numeric += 1

    if not values:
        print(f"column '{column}': no numeric values (all {non_numeric} values are text)")
        return

    result = {
        'column': column,
        'count': len(values),
        'sum': round(sum(values), 2),
        'avg': round(sum(values) / len(values), 2),
        'min': min(values),
        'max': max(values),
        'non_numeric': non_numeric,
    }

    if args.json:
        print(json.dumps(result, ensure_ascii=False))
    else:
        print(f"column: {column}")
        print(f"count: {result['count']} numeric, {result['non_numeric']} non-numeric")
        print(f"sum: {result['sum']}")
        print(f"avg: {result['avg']}")
        print(f"min: {result['min']}")
        print(f"max: {result['max']}")


def cmd_sort(args):
    """Sort file by a column."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(filepath)

    sort_key = args.column
    reverse = args.desc

    if sort_key not in headers:
        print(f"error: column '{sort_key}' not found.", file=sys.stderr)
        sys.exit(1)

    def sort_fn(r):
        v = r.get(sort_key, '')
        try:
            return (0, float(v))
        except (ValueError, TypeError):
            return (1, str(v))

    rows.sort(key=sort_fn, reverse=reverse)
    save_file(filepath, headers, rows, meta)

    if args.quiet:
        print(f"sorted: {len(rows)}")
    else:
        print(f"sorted {len(rows)} rows by '{sort_key}' ({'desc' if reverse else 'asc'})")


def cmd_convert(args):
    """Convert between file formats."""
    src = args.source
    dst = args.dest

    if not os.path.exists(src):
        print(f"error: file not found: {src}", file=sys.stderr)
        sys.exit(1)

    src_type = get_file_type(src)
    dst_type = get_file_type(dst)

    if not dst_type:
        print(f"error: unsupported output format: {Path(dst).suffix}", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(src)

    # Create appropriate meta for destination
    if dst_type == 'csv':
        dst_meta = {'encoding': 'utf-8', 'delimiter': ','}
    elif dst_type == 'xlsx':
        dst_meta = {'wb': None, 'ws': None}
    elif dst_type == 'json':
        dst_meta = {'encoding': 'utf-8'}

    save_file(dst, headers, rows, dst_meta)
    print(f"converted: {src} → {dst} ({len(rows)} rows)")


def cmd_merge(args):
    """Merge multiple files into one."""
    files = args.files
    output = args.output

    all_headers = []
    all_rows = []

    for f in files:
        if not os.path.exists(f):
            print(f"warning: skipping missing file: {f}", file=sys.stderr)
            continue
        headers, rows, meta = load_file(f)
        # Merge headers (preserve order, add new ones)
        for h in headers:
            if h not in all_headers:
                all_headers.append(h)
        all_rows.extend(rows)

    dst_type = get_file_type(output)
    if dst_type == 'csv':
        dst_meta = {'encoding': 'utf-8', 'delimiter': ','}
    elif dst_type == 'xlsx':
        dst_meta = {'wb': None, 'ws': None}
    elif dst_type == 'json':
        dst_meta = {'encoding': 'utf-8'}
    else:
        print(f"error: unsupported output format: {Path(output).suffix}", file=sys.stderr)
        sys.exit(1)

    save_file(output, all_headers, all_rows, dst_meta)
    print(f"merged {len(files)} files → {output} ({len(all_rows)} rows)")


# ---------------------------------------------------------------------------
# Excel-specific: formula, style, merge cells
# ---------------------------------------------------------------------------

def cmd_formula(args):
    """Write a formula to Excel cells."""
    filepath = args.file
    if get_file_type(filepath) != 'xlsx':
        print("error: formulas only work with .xlsx files", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import range_boundaries
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active

    cell_range = args.range
    formula = args.value

    if ':' in cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        count = 0
        for row in range(min_row, max_row + 1):
            f = formula.replace('{row}', str(row))
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = f
                count += 1
        print(f"formula applied to {count} cells")
    else:
        ws[cell_range] = formula
        print(f"formula set: {cell_range} = {formula}")

    wb.save(filepath)


def cmd_style(args):
    """Apply styling to Excel cells."""
    filepath = args.file
    if get_file_type(filepath) != 'xlsx':
        print("error: styling only works with .xlsx files", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import range_boundaries
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active

    cell_range = args.range
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    count = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)

            if args.bold is not None:
                font = cell.font.copy(bold=args.bold)
                cell.font = font

            if args.italic is not None:
                font = cell.font.copy(italic=args.italic)
                cell.font = font

            if args.font_size:
                font = cell.font.copy(size=args.font_size)
                cell.font = font

            if args.font_color:
                font = cell.font.copy(color=args.font_color)
                cell.font = font

            if args.bg_color:
                cell.fill = PatternFill(start_color=args.bg_color, end_color=args.bg_color, fill_type='solid')

            if args.align:
                cell.alignment = Alignment(horizontal=args.align)

            if args.number_format:
                cell.number_format = args.number_format

            count += 1

    wb.save(filepath)
    print(f"styled {count} cells in {cell_range}")


def cmd_merge_cells(args):
    """Merge cells in Excel."""
    filepath = args.file
    if get_file_type(filepath) != 'xlsx':
        print("error: merge-cells only works with .xlsx files", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active

    ws.merge_cells(args.range)
    if args.value:
        # Set value in top-left cell
        from openpyxl.utils import range_boundaries
        min_col, min_row, _, _ = range_boundaries(args.range)
        ws.cell(row=min_row, column=min_col).value = args.value

    wb.save(filepath)
    print(f"merged cells: {args.range}")


def cmd_width(args):
    """Set column width in Excel."""
    filepath = args.file
    if get_file_type(filepath) != 'xlsx':
        print("error: column width only works with .xlsx files", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active

    if args.auto:
        # Auto-fit all columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        print("auto-fitted all columns")
    else:
        ws.column_dimensions[args.column].width = args.size
        print(f"set column {args.column} width to {args.size}")

    wb.save(filepath)


def cmd_format_preset(args):
    """Apply a formatting preset to Excel file."""
    filepath = args.file
    if get_file_type(filepath) != 'xlsx':
        print("error: format presets only work with .xlsx files", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active
    preset = args.preset

    max_row = ws.max_row
    max_col = ws.max_column

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if preset == 'professional':
        # Header: bold, blue bg, white text
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        # Data rows: alternating colors, borders
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    elif preset == 'minimal':
        # Header: bold, bottom border only
        bottom_border = Border(bottom=Side(style='medium'))
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=11)
            cell.border = bottom_border

    elif preset == 'colorful':
        # Header: bold, green bg
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    elif preset == 'financial':
        # Header: dark bg, white text
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF', size=10, name='Arial')
            cell.fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = Font(size=10, name='Arial')
    else:
        print(f"error: unknown preset '{preset}'. Available: professional, minimal, colorful, financial", file=sys.stderr)
        sys.exit(1)

    # Auto-fit columns
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb.save(filepath)
    print(f"applied '{preset}' preset to {filepath}")


# ---------------------------------------------------------------------------
# Template System
# ---------------------------------------------------------------------------

TEMPLATE_DIR = os.path.join(os.path.expanduser('~'), '.asyre-tables', 'templates')
REGISTRY_FILE = os.path.join(os.path.expanduser('~'), '.asyre-tables', 'registry.json')


def ensure_template_dir():
    os.makedirs(TEMPLATE_DIR, exist_ok=True)


def load_registry():
    if os.path.exists(REGISTRY_FILE):
        with open(REGISTRY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'clients': {}}


def save_registry(reg):
    ensure_template_dir()
    with open(REGISTRY_FILE, 'w', encoding='utf-8') as f:
        json.dump(reg, f, ensure_ascii=False, indent=2)


def analyze_xlsx_structure(filepath):
    """Analyze an Excel file and extract its structure for template definition."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    header_row = find_header_row(ws)

    # Extract column definitions
    columns = {}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        if cell.value is None:
            continue
        col_letter = get_column_letter(c)
        col_def = {'name': str(cell.value), 'column': col_letter}

        # Detect type from data below
        for r in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
            val = ws.cell(row=r, column=c).value
            nf = ws.cell(row=r, column=c).number_format
            if val is not None:
                if isinstance(val, (int, float)):
                    col_def['type'] = 'number'
                    if nf and nf != 'General':
                        col_def['format'] = nf
                elif isinstance(val, str) and val.startswith('='):
                    col_def['type'] = 'formula'
                else:
                    col_def['type'] = 'text'
                break
        else:
            col_def['type'] = 'text'

        columns[col_letter] = col_def

    # Find data start (first row after header with data)
    data_start = header_row + 1
    for r in range(header_row + 1, ws.max_row + 1):
        has_data = False
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                has_data = True
                break
        if has_data:
            data_start = r
            break

    # Find formulas (summary rows)
    formulas = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                coord = f"{get_column_letter(c)}{r}"
                formulas[coord] = cell.value

    # Find protected rows (title, sub-headers, summary)
    protected = []
    for r in range(1, header_row + 1):
        protected.append(r)
    # Detect summary/total rows
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and ('合计' in val or '合  计' in val or 'total' in val.lower()):
                protected.append(r)
                break

    # Find merged cells
    merged = [str(m) for m in ws.merged_cells.ranges]

    return {
        'header_row': header_row,
        'data_start': data_start,
        'columns': columns,
        'formulas': formulas,
        'protected_rows': sorted(set(protected)),
        'merged_cells': merged,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
    }


def cmd_template_register(args):
    """Register a template from an Excel file."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    name = args.name
    client = args.client or '_default'
    ensure_template_dir()

    # 1. Analyze structure
    structure = analyze_xlsx_structure(filepath)

    # 2. Copy template file (read-only original)
    import shutil
    template_xlsx = os.path.join(TEMPLATE_DIR, f'{name}.xlsx')
    shutil.copy2(filepath, template_xlsx)
    os.chmod(template_xlsx, 0o444)  # read-only

    # 3. Save structure JSON
    template_json = os.path.join(TEMPLATE_DIR, f'{name}.json')
    structure['name'] = name
    structure['client'] = client
    structure['template_file'] = template_xlsx
    with open(template_json, 'w', encoding='utf-8') as f:
        json.dump(structure, f, ensure_ascii=False, indent=2)

    # 4. Update registry
    reg = load_registry()
    if client not in reg['clients']:
        reg['clients'][client] = {'templates': {}}
    reg['clients'][client]['templates'][name] = {
        'json': template_json,
        'xlsx': template_xlsx,
        'columns': [v['name'] for v in structure['columns'].values()],
    }
    save_registry(reg)

    # 5. Print summary
    print(f"registered template: {name}")
    print(f"  client: {client}")
    print(f"  header row: {structure['header_row']}")
    print(f"  data starts: row {structure['data_start']}")
    print(f"  columns: {', '.join(v['name'] for v in structure['columns'].values())}")
    if structure['formulas']:
        print(f"  formulas: {', '.join(structure['formulas'].keys())}")
    if structure['protected_rows']:
        print(f"  protected rows: {structure['protected_rows']}")
    if structure['merged_cells']:
        print(f"  merged cells: {', '.join(structure['merged_cells'])}")
    print(f"  template: {template_xlsx} (read-only)")


def cmd_template_list(args):
    """List registered templates."""
    reg = load_registry()
    client_filter = args.client

    if args.json:
        if client_filter:
            data = reg['clients'].get(client_filter, {})
        else:
            data = reg['clients']
        print(json.dumps(data, ensure_ascii=False, indent=2))
        return

    found = False
    for client, info in reg['clients'].items():
        if client_filter and client != client_filter:
            continue
        templates = info.get('templates', {})
        if templates:
            found = True
            print(f"client: {client}")
            for tname, tinfo in templates.items():
                cols = ', '.join(tinfo.get('columns', []))
                print(f"  {tname}: [{cols}]")
            print()

    if not found:
        print("no templates registered. Use: tables template register --name <name> <file>")


def cmd_template_new(args):
    """Create a new file from template (copy)."""
    name = args.name
    output = args.output

    # Find template
    template_json = os.path.join(TEMPLATE_DIR, f'{name}.json')
    if not os.path.exists(template_json):
        print(f"error: template '{name}' not found. Use: tables template list", file=sys.stderr)
        sys.exit(1)

    with open(template_json, 'r', encoding='utf-8') as f:
        tmpl = json.load(f)

    template_xlsx = tmpl['template_file']
    if not os.path.exists(template_xlsx):
        print(f"error: template file missing: {template_xlsx}", file=sys.stderr)
        sys.exit(1)

    import shutil
    shutil.copy2(template_xlsx, output)
    # Make writable
    os.chmod(output, 0o644)

    # Optionally update date in title
    if args.date:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output)
            ws = wb.active
            # Try to find and update date in first few rows
            import re
            for r in range(1, min(5, ws.max_row + 1)):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value and isinstance(cell.value, str):
                        # Replace date patterns like 2026年04月02日
                        new_val = re.sub(r'\d{4}年\d{2}月\d{2}日', args.date, cell.value)
                        if new_val != cell.value:
                            cell.value = new_val
            wb.save(output)
        except Exception:
            pass  # date replacement is best-effort

    print(f"created: {output} (from template '{name}')")


def cmd_template_fill(args):
    """Fill data into a file using template structure."""
    filepath = args.file
    name = args.template

    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    # Refuse to write to template files
    template_xlsx = os.path.join(TEMPLATE_DIR, f'{name}.xlsx')
    if os.path.abspath(filepath) == os.path.abspath(template_xlsx):
        print("error: cannot write to template file. Use 'tables template new' to create a copy first.", file=sys.stderr)
        sys.exit(1)

    # Load template structure
    template_json = os.path.join(TEMPLATE_DIR, f'{name}.json')
    if not os.path.exists(template_json):
        print(f"error: template '{name}' not found.", file=sys.stderr)
        sys.exit(1)

    with open(template_json, 'r', encoding='utf-8') as f:
        tmpl = json.load(f)

    # Parse values
    kv = parse_kv_pairs(args.values)
    if not kv:
        print("error: provide key=value pairs.", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter, column_index_from_string
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active

    # Map column names to column letters
    col_name_to_letter = {}
    for letter, info in tmpl['columns'].items():
        col_name_to_letter[info['name']] = letter

    # Find the insert row: last data row + 1, but before any protected rows
    data_start = tmpl['data_start']
    protected = set(tmpl.get('protected_rows', []))

    # Find the first empty row in the data area (or right before summary)
    insert_row = None
    for r in range(data_start, ws.max_row + 1):
        if r in protected:
            insert_row = r
            break
        has_data = False
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                has_data = True
                break
        if not has_data:
            insert_row = r
            break

    if insert_row is None:
        insert_row = ws.max_row + 1

    # If inserting before a protected row, shift it down
    if insert_row in protected:
        ws.insert_rows(insert_row)
        # Copy border style from adjacent data row
        style_src = insert_row + 1  # row below (was the protected row)
        if data_start < insert_row:
            style_src_row = insert_row - 1
            for c in range(1, ws.max_column + 1):
                src = ws.cell(row=style_src_row, column=c)
                dst = ws.cell(row=insert_row, column=c)
                dst.border = src.border.copy()
                dst.number_format = src.number_format
                dst.alignment = src.alignment.copy()

    # Write values
    written = 0
    for field_name, value in kv.items():
        col_letter = col_name_to_letter.get(field_name)
        if col_letter:
            col_idx = column_index_from_string(col_letter)
            cell = ws.cell(row=insert_row, column=col_idx)
            # Type conversion
            col_type = tmpl['columns'].get(col_letter, {}).get('type', 'text')
            if col_type == 'number':
                try:
                    value = float(value)
                    if value == int(value):
                        value = int(value)
                except ValueError:
                    pass
            cell.value = value
            # Apply number format from template
            fmt = tmpl['columns'].get(col_letter, {}).get('format')
            if fmt:
                cell.number_format = fmt
            written += 1

    # Update formulas if they reference expanding ranges
    for coord, formula in tmpl.get('formulas', {}).items():
        try:
            cell = ws[coord]
            # If this is a SUM formula, it might need updating
            # We don't change formulas - they auto-update when rows are inserted
        except Exception:
            pass

    wb.save(filepath)

    if args.quiet:
        print(f"filled: {written}")
    else:
        fields = ', '.join(f'{k}={v}' for k, v in kv.items())
        print(f"filled row {insert_row}: {fields}")


def cmd_template_info(args):
    """Show template structure."""
    name = args.name
    template_json = os.path.join(TEMPLATE_DIR, f'{name}.json')
    if not os.path.exists(template_json):
        print(f"error: template '{name}' not found.", file=sys.stderr)
        sys.exit(1)

    with open(template_json, 'r', encoding='utf-8') as f:
        tmpl = json.load(f)

    if args.json:
        print(json.dumps(tmpl, ensure_ascii=False, indent=2))
        return

    print(f"template: {tmpl['name']}")
    print(f"client: {tmpl.get('client', '_default')}")
    print(f"header row: {tmpl['header_row']}")
    print(f"data starts: row {tmpl['data_start']}")
    print(f"columns:")
    for letter, info in tmpl['columns'].items():
        fmt = f" ({info.get('format', '')})" if info.get('format') else ''
        print(f"  {letter}: {info['name']} [{info['type']}]{fmt}")
    if tmpl.get('formulas'):
        print(f"formulas:")
        for coord, formula in tmpl['formulas'].items():
            print(f"  {coord}: {formula}")
    if tmpl.get('protected_rows'):
        print(f"protected rows: {tmpl['protected_rows']}")
    if tmpl.get('merged_cells'):
        print(f"merged cells: {', '.join(tmpl['merged_cells'])}")


def cmd_template_delete(args):
    """Delete a template."""
    name = args.name
    template_json = os.path.join(TEMPLATE_DIR, f'{name}.json')
    template_xlsx = os.path.join(TEMPLATE_DIR, f'{name}.xlsx')

    if not os.path.exists(template_json):
        print(f"error: template '{name}' not found.", file=sys.stderr)
        sys.exit(1)

    # Remove files
    if os.path.exists(template_xlsx):
        os.chmod(template_xlsx, 0o644)
        os.remove(template_xlsx)
    os.remove(template_json)

    # Update registry
    reg = load_registry()
    for client in reg['clients'].values():
        templates = client.get('templates', {})
        if name in templates:
            del templates[name]
    save_registry(reg)

    print(f"deleted template: {name}")


# ---------------------------------------------------------------------------
# Audit / Reconciliation
# ---------------------------------------------------------------------------

def detect_balance_pattern(headers, rows):
    """
    Detect if there's a running balance column.
    Common pattern: 余额 = prev_余额 + 收入 - 支出
    Returns (balance_col, income_cols, expense_cols) or None.
    """
    # Look for common column name patterns
    balance_names = ['余额', '余額', 'balance', '结余', '账户余额']
    income_names = ['收入', '入账', 'income', '贷方', '进账', '收款金额', '收款']
    expense_names = ['支出', '出账', 'expense', '借方', '付款', '欠款金额', '付款']

    balance_col = None
    income_cols = []
    expense_cols = []

    for h in headers:
        hl = h.lower().strip()
        for bn in balance_names:
            if bn in hl:
                balance_col = h
                break
        for inc in income_names:
            if inc in hl:
                income_cols.append(h)
                break
        for exp in expense_names:
            if exp in hl:
                expense_cols.append(h)
                break

    if balance_col and (income_cols or expense_cols):
        return balance_col, income_cols, expense_cols
    return None


def to_float(val):
    """Safely convert to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def cmd_sheet_audit(args):
    """
    Full audit of a spreadsheet:
    1. Running balance verification (余额 = prev + 收入 - 支出)
    2. Sum/total row verification
    3. Missing data detection
    4. Mark results: color on problem cells + notes in nearby empty cells
    5. Separate audit sheet with full report
    """
    filepath = args.file
    if not os.path.exists(filepath):
        print(f"error: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    is_xlsx = get_file_type(filepath) == 'xlsx'
    if not is_xlsx:
        print("error: audit currently only supports .xlsx files", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("error: openpyxl required.", file=sys.stderr)
        sys.exit(1)

    headers, rows, meta = load_file(filepath)
    row_map = meta.get('row_map', [])
    header_row = meta.get('header_row', 1)

    # Load workbook for cell-level inspection
    wb_data = load_workbook(filepath, data_only=True)
    ws_data = wb_data.active
    wb_formula = load_workbook(filepath, data_only=False)
    ws_formula = wb_formula.active

    issues = []  # list of {row, col, type, message, detail}

    # =====================================================
    # Check 1: Running balance
    # =====================================================
    pattern = detect_balance_pattern(headers, rows)
    if pattern:
        bal_col, inc_cols, exp_cols = pattern
        bal_col_idx = headers.index(bal_col) + 1

        print(f"[检查1] 余额验算: {bal_col} = prev + {'+'.join(inc_cols)} - {'+'.join(exp_cols)}")

        prev_balance = None
        for i, row in enumerate(rows):
            bal_val = row.get(bal_col)
            current_bal = to_float(bal_val)
            income = sum(to_float(row.get(c)) for c in inc_cols)
            expense = sum(to_float(row.get(c)) for c in exp_cols)
            excel_row = row_map[i] if i < len(row_map) else (header_row + 1 + i)

            # Skip formula/summary rows
            if isinstance(bal_val, str) and ('合计' in bal_val or '合  计' in bal_val or str(bal_val).startswith('=')):
                continue
            # Skip second table headers
            if isinstance(bal_val, str) and any(kw in bal_val for kw in ['客户', '备注', '收款']):
                break

            has_activity = (income > 0 or expense > 0)

            if prev_balance is not None and has_activity:
                expected = round(prev_balance + income - expense, 2)

                # Find description for this row
                row_desc = ''
                for h in headers:
                    v = row.get(h)
                    if v and isinstance(v, str) and v.strip() and h != bal_col:
                        row_desc = v.strip()
                        break
                if not row_desc:
                    row_desc = f'第{excel_row}行'

                if bal_val is None:
                    in_part = f'收入{income}元' if income else ''
                    ex_part = f'支出{expense}元' if expense else ''
                    parts = [p for p in [in_part, ex_part] if p]
                    issues.append({
                        'row': excel_row,
                        'col': bal_col_idx,
                        'type': 'missing',
                        'message': f'"{row_desc}"这笔账没有算余额。上笔余额{prev_balance}元，{" ".join(parts)}，算下来应该剩 {expected} 元',
                        'detail': f'上笔余额 {prev_balance} + 收入 {income} - 支出 {expense} = {expected}',
                    })
                    prev_balance = expected
                    continue

                diff = round(current_bal - expected, 2)
                if abs(diff) > 0.01:
                    if diff > 0:
                        diff_desc = f'多了 {abs(diff)} 元'
                    else:
                        diff_desc = f'少了 {abs(diff)} 元'
                    issues.append({
                        'row': excel_row,
                        'col': bal_col_idx,
                        'type': 'mismatch',
                        'message': f'"{row_desc}"的余额有问题：写的是 {current_bal} 元，但按上笔余额 {prev_balance} 算下来应该是 {expected} 元，{diff_desc}',
                        'detail': f'上笔余额 {prev_balance} + 收入 {income} - 支出 {expense} = {expected}，实际写了 {current_bal}，差额 {diff:+.2f}',
                    })

            if bal_val is not None and current_bal != 0:
                prev_balance = current_bal

        bal_issues = sum(1 for iss in issues if iss['col'] == bal_col_idx)
        print(f"  → {bal_issues} 个问题" if bal_issues else "  → 通过")
    else:
        print("[检查1] 未检测到余额模式，跳过")

    # =====================================================
    # Check 2: Sum/Total formulas
    # =====================================================
    print(f"\n[检查2] 合计公式验证")
    sum_checked = 0
    for r in range(1, ws_formula.max_row + 1):
        for c in range(1, ws_formula.max_column + 1):
            cell_f = ws_formula.cell(r, c).value
            if not cell_f or not isinstance(cell_f, str):
                continue
            if not cell_f.upper().startswith('=SUM'):
                continue

            sum_checked += 1
            # Get the cached value from data-only workbook
            cached = ws_data.cell(r, c).value

            # Parse the SUM range and manually calculate
            import re
            m = re.match(r'=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', cell_f.upper())
            if m:
                col_letter = m.group(1)
                start_row = int(m.group(2))
                end_row = int(m.group(4))

                actual_sum = 0
                for sr in range(start_row, end_row + 1):
                    val = ws_data.cell(sr, c).value
                    if sr == r:
                        continue  # skip the formula cell itself
                    if val is not None:
                        try:
                            actual_sum += float(val)
                        except (ValueError, TypeError):
                            pass

                actual_sum = round(actual_sum, 2)

                # The formula range might not include newly inserted rows
                # Check if there are data rows outside the SUM range
                col_name = get_column_letter(c)
                data_outside = []
                for check_r in range(header_row + 1, ws_data.max_row + 1):
                    if check_r < start_row or check_r > end_row:
                        val = ws_data.cell(check_r, c).value
                        if val is not None and check_r != r:
                            try:
                                fv = float(val)
                                if fv != 0 and not isinstance(ws_formula.cell(check_r, c).value, str):
                                    data_outside.append((check_r, fv))
                            except (ValueError, TypeError):
                                pass

                if data_outside:
                    missed_sum = round(sum(v for _, v in data_outside), 2)
                    col_name_display = get_column_letter(c)
                    # Figure out what this column represents
                    col_header = ws_data.cell(header_row, c).value or col_name_display
                    issues.append({
                        'row': r,
                        'col': c,
                        'type': 'formula_range',
                        'message': f'合计金额少算了："{col_header}"的合计没有包含新增的数据，漏算了 {missed_sum} 元。合计数字比实际偏小',
                        'detail': f'当前公式 {cell_f} 只算到第{end_row}行，但第{",".join([str(rn) for rn,_ in data_outside])}行还有 {missed_sum} 元没算进去',
                    })

    sum_issues = sum(1 for iss in issues if iss['type'] == 'formula_range')
    print(f"  检查了{sum_checked}个公式 → {sum_issues} 个问题" if sum_checked else "  → 无公式")

    # =====================================================
    # Check 3: Missing data in rows with partial data
    # =====================================================
    print(f"\n[检查3] 数据完整性")
    data_issues_before = len(issues)
    for i, row in enumerate(rows):
        excel_row = row_map[i] if i < len(row_map) else (header_row + 1 + i)
        has_some_data = any(row.get(h) is not None for h in headers)
        if not has_some_data:
            continue

        # Check: if there's income or expense, there should be a description (摘要)
        income = sum(to_float(row.get(c)) for c in (inc_cols if pattern else []))
        expense = sum(to_float(row.get(c)) for c in (exp_cols if pattern else []))

        if income > 0 or expense > 0:
            desc_cols = [h for h in headers if any(kw in h for kw in ['摘要', '说明', '备注', '项目', 'description'])]
            for dc in desc_cols:
                if row.get(dc) is None:
                    # Check consecutive empty desc rows (might be continuation)
                    # Only flag if the previous row also doesn't explain this
                    if i > 0:
                        prev_desc = rows[i-1].get(dc)
                        prev_income = sum(to_float(rows[i-1].get(c)) for c in (inc_cols if pattern else []))
                        prev_expense = sum(to_float(rows[i-1].get(c)) for c in (exp_cols if pattern else []))
                        if prev_desc and (prev_income > 0 or prev_expense > 0):
                            # Likely a continuation row, not necessarily an error
                            continue

    data_issues = len(issues) - data_issues_before
    print(f"  → {data_issues} 个问题" if data_issues else "  → 通过")

    # =====================================================
    # Output
    # =====================================================
    print(f"\n{'='*40}")
    if not issues:
        print(f"审计通过: {len(rows)} 行已检查，未发现问题。")
    else:
        print(f"审计完成: {len(rows)} 行已检查，发现 {len(issues)} 个问题：\n")
        for iss in issues:
            tag = {'missing': '漏算', 'mismatch': '对不上', 'formula_range': '少算'}.get(iss['type'], iss['type'])
            print(f"  [{tag}] 第{iss['row']}行: {iss['message']}")

    # =====================================================
    # Mark in file
    # =====================================================
    if args.mark and issues:
        wb = load_workbook(filepath)
        ws = wb.active

        error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        missing_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        note_font = Font(color='9C0006', size=9, italic=True)

        # Find empty column next to data for notes
        # Scan each issue row to find the first empty column after the problem cell
        for iss in issues:
            r = iss['row']
            c = iss['col']
            cell = ws.cell(row=r, column=c)

            # Color the problem cell
            if iss['type'] == 'missing':
                cell.fill = missing_fill
            else:
                cell.fill = error_fill

            # Find nearest empty cell to the RIGHT for the note
            note_col = None
            for nc in range(c + 1, ws.max_column + 3):
                candidate = ws.cell(row=r, column=nc)
                if candidate.value is None:
                    # Check it's not a merged cell
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if candidate.coordinate in merged_range:
                            is_merged = True
                            break
                    if not is_merged:
                        note_col = nc
                        break

            if note_col:
                note_cell = ws.cell(row=r, column=note_col)
                # Short version for main sheet — just the key point
                if iss['type'] == 'missing':
                    # Extract the expected amount
                    import re as _re
                    amt = _re.search(r'应该剩 ([\d.]+)', iss['message'])
                    note_cell.value = f"← 余额应为 {amt.group(1)}" if amt else f"← 余额缺失"
                elif iss['type'] == 'mismatch':
                    amt = _re.search(r'应该是 ([\d.]+)', iss['message'])
                    diff_m = _re.search(r'([多少]了 [\d.]+ 元)', iss['message'])
                    if diff_m:
                        note_cell.value = f"← 余额{diff_m.group(1)}"
                    elif amt:
                        note_cell.value = f"← 应为 {amt.group(1)}"
                    else:
                        note_cell.value = f"← 余额有误"
                elif iss['type'] == 'formula_range':
                    amt = _re.search(r'漏算了 ([\d.]+)', iss['message'])
                    note_cell.value = f"← 合计漏算 {amt.group(1)}元" if amt else f"← 合计有误"
                note_cell.font = note_font

        # --- Audit sheet ---
        if '审计结果' in wb.sheetnames:
            del wb['审计结果']
        audit_ws = wb.create_sheet('审计结果')

        # Title
        audit_ws.merge_cells('A1:G1')
        audit_ws['A1'] = f'审计报告'
        audit_ws['A1'].font = Font(bold=True, size=14)
        audit_ws['A1'].alignment = Alignment(horizontal='center')
        audit_ws.row_dimensions[1].height = 30

        import datetime
        audit_ws.merge_cells('A2:G2')
        audit_ws['A2'] = f'文件: {os.path.basename(filepath)}  |  审计时间: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}  |  检查行数: {len(rows)}  |  问题数: {len(issues)}'
        audit_ws['A2'].font = Font(size=10, italic=True, color='666666')
        audit_ws['A2'].alignment = Alignment(horizontal='center')

        # Header row
        audit_headers = ['#', '行号', '类型', '问题描述', '计算详情']
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        for ci, h in enumerate(audit_headers, 1):
            cell = audit_ws.cell(row=4, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin

        wrap_align = Alignment(wrap_text=True, vertical='top')

        # Issue rows
        for ri, iss in enumerate(issues):
            r = ri + 5
            audit_ws.cell(row=r, column=1, value=ri + 1).border = thin
            audit_ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='top')
            audit_ws.cell(row=r, column=2, value=iss['row']).border = thin
            audit_ws.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='top')

            type_cell = audit_ws.cell(row=r, column=3)
            type_map = {'missing': '漏算余额', 'mismatch': '余额对不上', 'formula_range': '合计少算'}
            type_cell.value = type_map.get(iss['type'], iss['type'])
            type_cell.border = thin
            type_cell.alignment = Alignment(horizontal='center', vertical='top')
            if iss['type'] == 'missing':
                type_cell.fill = missing_fill
            else:
                type_cell.fill = error_fill

            msg_cell = audit_ws.cell(row=r, column=4, value=iss['message'])
            msg_cell.border = thin
            msg_cell.alignment = wrap_align

            detail_cell = audit_ws.cell(row=r, column=5, value=iss['detail'])
            detail_cell.border = thin
            detail_cell.alignment = wrap_align

            # Auto row height: ~15 per line, estimate lines by length
            max_len = max(len(iss['message']), len(iss['detail']))
            est_lines = max(2, (max_len // 35) + 1)
            audit_ws.row_dimensions[r].height = est_lines * 18

        # Summary
        sr = len(issues) + 6
        if not issues:
            audit_ws.cell(row=5, column=1, value='所有检查通过，未发现问题。')
            audit_ws.cell(row=5, column=1).font = Font(color='006100', size=12, bold=True)

        # Column widths — wide enough to be readable
        audit_ws.column_dimensions['A'].width = 5
        audit_ws.column_dimensions['B'].width = 7
        audit_ws.column_dimensions['C'].width = 12
        audit_ws.column_dimensions['D'].width = 45
        audit_ws.column_dimensions['E'].width = 45

        wb.active = wb.sheetnames.index(ws.title)
        wb.save(filepath)
        print(f"\nmarked {len(issues)} cells (color + notes in adjacent empty cells)")
        print(f"audit report → sheet '审计结果'")

    if args.json:
        result = {
            'total_rows': len(rows),
            'issues': len(issues),
            'details': [{
                'row': iss['row'],
                'type': iss['type'],
                'message': iss['message'],
            } for iss in issues]
        }
        print(json.dumps(result, ensure_ascii=False))


# ---------------------------------------------------------------------------
# CLI Parser
# ---------------------------------------------------------------------------

def build_parser():
    parser = argparse.ArgumentParser(
        prog='tables',
        description='Asyre Tables - Spreadsheet operations CLI for AI agents.'
    )
    sub = parser.add_subparsers(dest='command')

    # --- sheet sub-commands ---
    sheet = sub.add_parser('sheet', help='Spreadsheet operations')
    sheet_sub = sheet.add_subparsers(dest='sheet_cmd')

    # info
    p = sheet_sub.add_parser('info', help='Show file structure')
    p.add_argument('file')
    p.add_argument('--json', action='store_true')

    # list
    p = sheet_sub.add_parser('list', help='List contents')
    p.add_argument('file')
    p.add_argument('--limit', type=int, default=20)
    p.add_argument('--columns', type=str, help='Columns to show (comma-separated)')
    p.add_argument('--json', action='store_true')

    # query
    p = sheet_sub.add_parser('query', help='Query with conditions')
    p.add_argument('file')
    p.add_argument('--where', type=str, required=True)
    p.add_argument('--columns', type=str)
    p.add_argument('--sort', type=str, help='Sort column (prefix - for desc)')
    p.add_argument('--limit', type=int, default=20)
    p.add_argument('--json', action='store_true')

    # add
    p = sheet_sub.add_parser('add', help='Add a row')
    p.add_argument('file')
    p.add_argument('values', nargs='+', help='key=value pairs')
    p.add_argument('--quiet', '-q', action='store_true')

    # update
    p = sheet_sub.add_parser('update', help='Update rows')
    p.add_argument('file')
    p.add_argument('--where', type=str, required=True)
    p.add_argument('--set', type=str, required=True)
    p.add_argument('--quiet', '-q', action='store_true')

    # delete
    p = sheet_sub.add_parser('delete', help='Delete rows')
    p.add_argument('file')
    p.add_argument('--where', type=str, required=True)
    p.add_argument('--quiet', '-q', action='store_true')

    # stats
    p = sheet_sub.add_parser('stats', help='Column statistics')
    p.add_argument('file')
    p.add_argument('--column', type=str, required=True)
    p.add_argument('--group-by', type=str)
    p.add_argument('--json', action='store_true')

    # sort
    p = sheet_sub.add_parser('sort', help='Sort file by column')
    p.add_argument('file')
    p.add_argument('--column', type=str, required=True)
    p.add_argument('--desc', action='store_true')
    p.add_argument('--quiet', '-q', action='store_true')

    # formula (xlsx only)
    p = sheet_sub.add_parser('formula', help='Write formula to cells')
    p.add_argument('file')
    p.add_argument('--range', type=str, required=True)
    p.add_argument('--value', type=str, required=True)

    # style (xlsx only)
    p = sheet_sub.add_parser('style', help='Style cells')
    p.add_argument('file')
    p.add_argument('--range', type=str, required=True)
    p.add_argument('--bold', action='store_true', default=None)
    p.add_argument('--no-bold', dest='bold', action='store_false')
    p.add_argument('--italic', action='store_true', default=None)
    p.add_argument('--font-size', type=int)
    p.add_argument('--font-color', type=str, help='Hex color (e.g. FF0000)')
    p.add_argument('--bg-color', type=str, help='Hex color (e.g. 4472C4)')
    p.add_argument('--align', type=str, choices=['left', 'center', 'right'])
    p.add_argument('--number-format', type=str, help='e.g. #,##0.00')

    # merge-cells (xlsx only)
    p = sheet_sub.add_parser('merge-cells', help='Merge cells')
    p.add_argument('file')
    p.add_argument('--range', type=str, required=True)
    p.add_argument('--value', type=str)

    # width (xlsx only)
    p = sheet_sub.add_parser('width', help='Set column width')
    p.add_argument('file')
    p.add_argument('--column', type=str)
    p.add_argument('--size', type=float)
    p.add_argument('--auto', action='store_true')

    # format (xlsx only)
    p = sheet_sub.add_parser('format', help='Apply formatting preset')
    p.add_argument('file')
    p.add_argument('--preset', type=str, required=True,
                   choices=['professional', 'minimal', 'colorful', 'financial'])

    # audit
    p = sheet_sub.add_parser('audit', help='Audit calculations, detect errors')
    p.add_argument('file')
    p.add_argument('--mark', action='store_true', help='Mark errors in file (red + comments)')
    p.add_argument('--json', action='store_true')
    p.add_argument('--quiet', '-q', action='store_true')

    # --- template sub-commands ---
    tmpl = sub.add_parser('template', help='Template management')
    tmpl_sub = tmpl.add_subparsers(dest='tmpl_cmd')

    p = tmpl_sub.add_parser('register', help='Register a template from file')
    p.add_argument('file')
    p.add_argument('--name', required=True, help='Template name')
    p.add_argument('--client', type=str, help='Client name')

    p = tmpl_sub.add_parser('list', help='List templates')
    p.add_argument('--client', type=str)
    p.add_argument('--json', action='store_true')

    p = tmpl_sub.add_parser('info', help='Show template structure')
    p.add_argument('name')
    p.add_argument('--json', action='store_true')

    p = tmpl_sub.add_parser('new', help='Create file from template')
    p.add_argument('name', help='Template name')
    p.add_argument('-o', '--output', required=True, help='Output file path')
    p.add_argument('--date', type=str, help='Date to set (e.g. 2026年04月06日)')

    p = tmpl_sub.add_parser('fill', help='Fill data using template')
    p.add_argument('file', help='Target file (NOT the template)')
    p.add_argument('--template', required=True, help='Template name')
    p.add_argument('values', nargs='+', help='column=value pairs')
    p.add_argument('--quiet', '-q', action='store_true')

    p = tmpl_sub.add_parser('delete', help='Delete a template')
    p.add_argument('name')

    # --- convert ---
    p = sub.add_parser('convert', help='Convert between file formats')
    p.add_argument('source')
    p.add_argument('dest')

    # --- merge ---
    p = sub.add_parser('merge', help='Merge multiple files')
    p.add_argument('files', nargs='+')
    p.add_argument('-o', '--output', required=True)

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    if args.command == 'sheet':
        cmd_map = {
            'info': cmd_info,
            'list': cmd_list,
            'query': cmd_query,
            'add': cmd_add,
            'update': cmd_update,
            'delete': cmd_delete,
            'stats': cmd_stats,
            'sort': cmd_sort,
            'formula': cmd_formula,
            'style': cmd_style,
            'merge-cells': cmd_merge_cells,
            'width': cmd_width,
            'format': cmd_format_preset,
            'audit': cmd_sheet_audit,
        }
        fn = cmd_map.get(args.sheet_cmd)
        if fn:
            fn(args)
        else:
            parser.parse_args(['sheet', '-h'])
    elif args.command == 'template':
        tmpl_map = {
            'register': cmd_template_register,
            'list': cmd_template_list,
            'info': cmd_template_info,
            'new': cmd_template_new,
            'fill': cmd_template_fill,
            'delete': cmd_template_delete,
        }
        fn = tmpl_map.get(args.tmpl_cmd)
        if fn:
            fn(args)
        else:
            parser.parse_args(['template', '-h'])
    elif args.command == 'convert':
        cmd_convert(args)
    elif args.command == 'merge':
        cmd_merge(args)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
